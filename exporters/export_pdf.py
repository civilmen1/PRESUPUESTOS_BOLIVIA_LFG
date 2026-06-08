"""Exportación de los Formularios (B-1 a B-5, A-8, A-9) a PDF.

Genera el PDF reproduciendo EXACTAMENTE el formato del Excel de formularios:
lee el libro generado por export_formularios y vuelca cada hoja como tabla.

- Las filas que son texto fuera de tabla (títulos, datos generales, notas) se
  muestran como párrafos SIN ajuste de texto (una sola línea, recortada si es
  necesaria al ancho de página).
- Las tablas de recursos conservan columnas, bordes y rellenos del Excel.
"""
from __future__ import annotations

from pathlib import Path

from config import settings
from config.logging_config import get_logger
from core import repositories

logger = get_logger(__name__)


def exportar_pdf(proyecto_id: int, ruta: str | Path | None = None) -> Path:
    proyecto = repositories.obtener_proyecto(proyecto_id)
    if not proyecto:
        raise ValueError(f"Proyecto {proyecto_id} no encontrado")
    if ruta is None:
        ruta = settings.EXPORT_DIR / f"formularios_proyecto_{proyecto_id}.pdf"
    ruta = Path(ruta)
    try:
        return _exportar_desde_excel(proyecto_id, ruta)
    except ImportError:
        logger.warning("reportlab no instalado; exportando a texto plano")
        return _exportar_txt(proyecto, proyecto_id, ruta.with_suffix(".txt"))


def _exportar_desde_excel(proyecto_id: int, ruta: Path) -> Path:
    """Genera el Excel de formularios y lo reproduce en PDF, hoja por hoja."""
    from openpyxl import load_workbook
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    from exporters.export_formularios import exportar_formularios
    xlsx = exportar_formularios(proyecto_id)
    wb = load_workbook(xlsx, data_only=True)

    # Estilos de texto SIN ajuste de linea (una sola linea por celda de texto).
    est_titulo = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=11,
                                leading=13, wordWrap=None)
    est_sub = ParagraphStyle("s", fontName="Helvetica-Bold", fontSize=9,
                             leading=11, wordWrap=None)
    est_norm = ParagraphStyle("n", fontName="Helvetica", fontSize=8, leading=10,
                              wordWrap=None)

    doc = SimpleDocTemplate(str(ruta), pagesize=landscape(A4),
                            topMargin=1 * cm, bottomMargin=1 * cm,
                            leftMargin=1 * cm, rightMargin=1 * cm)
    story = []
    ancho_util = landscape(A4)[0] - 2 * cm

    for hi, nombre in enumerate(wb.sheetnames):
        ws = wb[nombre]
        if hi > 0:
            story.append(PageBreak())

        filas = list(ws.iter_rows(values_only=True))
        ncols = max((len([c for c in r if c is not None]) for r in filas
                     if r), default=1)
        # detectar el bloque de tabla (filas con >=3 celdas) vs texto suelto
        bloque_tabla: list[list] = []

        def _volcar_tabla():
            if not bloque_tabla:
                return
            maxc = max(len(r) for r in bloque_tabla)
            data = [[("" if c is None else str(c)) for c in
                     (r + [""] * (maxc - len(r)))] for r in bloque_tabla]
            colw = ancho_util / maxc
            tabla = Table(data, colWidths=[colw] * maxc)
            tabla.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1),
                 [colors.white, colors.HexColor("#f4f6f9")]),
            ]))
            story.append(tabla)
            story.append(Spacer(1, 0.2 * cm))
            bloque_tabla.clear()

        for r in filas:
            valores = [c for c in (r or []) if c is not None and str(c).strip()]
            n_valores = len(valores)
            if n_valores >= 3:
                # fila de tabla
                bloque_tabla.append(list(r))
            else:
                _volcar_tabla()
                if n_valores == 0:
                    continue
                # texto fuera de tabla: una sola linea, sin ajuste
                texto = "   ".join(str(c) for c in valores)
                texto = _escapar(texto)
                if "FORMULARIO" in texto.upper():
                    story.append(Paragraph(texto, est_titulo))
                elif texto.isupper() and len(texto) < 60:
                    story.append(Paragraph(texto, est_sub))
                else:
                    story.append(Paragraph(_recortar(texto, 150), est_norm))
        _volcar_tabla()

    doc.build(story)
    logger.info("PDF de formularios generado: %s", ruta)
    return ruta


def _escapar(t: str) -> str:
    return t.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _recortar(t: str, n: int) -> str:
    return t if len(t) <= n else t[: n - 1] + "…"


def _exportar_txt(proyecto, proyecto_id, ruta: Path) -> Path:
    lineas = [f"Formularios - {proyecto.nombre} ({proyecto.moneda})", "=" * 60]
    for item in repositories.listar_items(proyecto_id):
        res = repositories.obtener_resultado(item.id)
        lineas.append(f"\n{item.numero or ''} {item.descripcion} [{item.unidad}]")
        for r in repositories.listar_recursos(item.id):
            lineas.append(f"  - {r.tipo}: {r.descripcion} {r.cantidad_apu:g} "
                          f"{r.unidad} x {r.precio_unitario:,.2f} = {r.subtotal:,.2f}")
        if res:
            lineas.append(f"  P.U. Total: {res.precio_unitario_total:,.2f}")
    ruta.write_text("\n".join(lineas), encoding="utf-8")
    return ruta
