"""Exportación de los Formularios oficiales NB-SABS (DS 0181) a Excel.

  B-1  Presupuesto General de la Obra (por ítems y módulos)
  B-2  Análisis de Precios Unitarios — UNA PÁGINA POR ÍTEM, con fórmulas vivas
  B-3  Precios Unitarios de Elementales (materiales, mano de obra, equipo)
  B-4  Equipo Mínimo Comprometido
  A-8  Cronograma de Ejecución de Obra
  A-9  Cronograma de Movilización de Equipo
  B-5  Cronograma de Desembolsos

Características:
  - Cada formulario lleva fórmulas reales de Excel (no solo valores).
  - Pie de página "Página 1 de ?" en cada hoja (al imprimir).
  - Textos ajustados dentro de las celdas (wrap + alto de fila automático).
  - Precisión de pantalla activada (precision_as_displayed) para que los
    cálculos usen los valores tal como se ven (2 decimales).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import settings
from core import repositories
from core.sabs import calcular_desglose
from models.apu_resource import TIPO_EQUIPO, TIPO_MANO_OBRA, TIPO_MATERIAL

# ---------------------------------------------------------------- estilos
_AZUL = "1F4E78"
_GRIS = "D9D9D9"
_BORDE = Border(*[Side(style="thin", color="999999")] * 4)
_F_TITULO = Font(bold=True, size=14, color="1F4E78")
_F_HEAD = Font(bold=True, color="FFFFFF", size=9)
_F_BOLD = Font(bold=True, size=9)
_F_NORM = Font(size=9)
_FILL_HEAD = PatternFill("solid", fgColor=_AZUL)
_FILL_GRIS = PatternFill("solid", fgColor=_GRIS)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
# Variantes SIN ajuste de texto (wrap_text=False) para el texto que va FUERA de
# las tablas: encabezado (títulos y datos del proyecto), notas, subtítulos y pie
# de firma. Dentro de las tablas se mantiene el wrap (descripciones largas).
_CENTER_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
_RIGHT_NW = Alignment(horizontal="right", vertical="center", wrap_text=False)
_LEFT_NW = Alignment(horizontal="left", vertical="center", wrap_text=False)

_FMT_MONTO = "#,##0.00"
_FMT_CANT = "#,##0.0000"

# Estado de moneda del libro en exportación (se fija en exportar_formularios).
# Los montos se calculan en BOB; se convierten a la moneda del proyecto al
# escribir los VALORES de entrada (precios), de modo que las fórmulas vivas de
# Excel queden internamente consistentes en la moneda mostrada.
_MONEDA = {"sim": "Bs", "factor": 1.0}


def _money(valor_bob: float) -> float:
    """Convierte un monto en BOB a la moneda activa del libro."""
    try:
        return round(float(valor_bob) * _MONEDA["factor"], 2)
    except (TypeError, ValueError):
        return valor_bob


def _sim() -> str:
    """Símbolo de la moneda activa (Bs o $us)."""
    return _MONEDA["sim"]


def _set(ws, celda, valor, font=_F_NORM, align=None, fill=None, fmt=None):
    c = ws[celda]
    c.value = valor
    c.font = font
    c.alignment = align or _LEFT  # wrap por defecto en todas las celdas
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


def _borde_rango(ws, fila, cols):
    for c in cols:
        ws[f"{c}{fila}"].border = _BORDE


def _config_pagina(ws, ncols, titulo_cols=6):
    """Ajusta la hoja para impresión: A4 horizontal/vertical, ajustar a ancho,
    pie 'Página X de N', y márgenes. El alto de filas se autoajusta al wrap."""
    ws.page_setup.orientation = ("landscape" if ncols > 6 else "portrait")
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddFooter.center.text = "Página &P de &N"
    ws.oddFooter.right.text = "&D"
    # repetir el bloque de encabezado en cada página impresa
    ws.print_title_rows = "1:4"
    for m in ("left", "right", "top", "bottom"):
        setattr(ws.page_margins, m, 0.5)


def _autoalto_filas(ws, fila_ini, fila_fin, col_texto="B", ancho_col=44):
    """Estima el alto de fila según el largo del texto para que el wrap se vea."""
    for fila in range(fila_ini, fila_fin + 1):
        celda = ws[f"{col_texto}{fila}"]
        texto = str(celda.value or "")
        if texto:
            # ~ caracteres por línea según ancho de columna
            chars_linea = max(int(ancho_col * 1.1), 10)
            n_lineas = max(1, -(-len(texto) // chars_linea))
            ws.row_dimensions[fila].height = max(15, n_lineas * 12.5)


def _encabezado(ws, codigo: str, titulo: str, proyecto, ncols: int = 6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _set(ws, "A1", f"FORMULARIO {codigo}", _F_TITULO, _CENTER_NW)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    _set(ws, "A2", titulo.upper(), _F_BOLD, _CENTER_NW)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    tc_txt = (f"   |   T/C: {proyecto.tipo_cambio:.2f} Bs/$us"
              if proyecto.moneda == "USD" else "")
    _set(ws, "A3", f"Proyecto: {proyecto.nombre}   |   Entidad: "
                   f"{proyecto.entidad or '—'}   |   Moneda: {proyecto.moneda}"
                   f"{tc_txt}", _F_NORM, _CENTER_NW)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
    _set(ws, "A4", f"Proponente: {proyecto.proponente or '—'}   |   "
                   f"Departamento: {proyecto.region or '—'}", _F_NORM, _CENTER_NW)
    return 6  # primera fila de contenido


def _fila_headers(ws, fila, headers, anchos):
    for i, (h, w) in enumerate(zip(headers, anchos), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        _set(ws, f"{col}{fila}", h, _F_HEAD, _CENTER, _FILL_HEAD)
        ws[f"{col}{fila}"].border = _BORDE


# ===================================================================== B-1
def _b1_presupuesto(ws, proyecto, items):
    s = _sim()
    fila = _encabezado(ws, "B-1", "Presupuesto General de la Obra", proyecto, 6)
    _fila_headers(ws, fila, ["N°", "Ítem / Descripción", "Unidad", "Cantidad",
                             f"Precio Unitario ({s})", f"Precio Total ({s})"],
                  [6, 50, 10, 12, 16, 18])
    fila += 1
    ini_texto = fila
    total_general = 0.0
    filas_total = []          # filas con fórmula de total para el SUM final
    modulo_actual = None
    for it in items:
        mod = _modulo_nombre(it)
        if mod and mod != modulo_actual:
            modulo_actual = mod
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila,
                           end_column=6)
            _set(ws, f"A{fila}", mod, _F_BOLD, _LEFT, _FILL_GRIS)
            _borde_rango(ws, fila, "ABCDEF")
            fila += 1
        res = repositories.obtener_resultado(it.id)
        pu = res.precio_unitario_total if res else 0.0
        total_general += pu * it.cantidad
        _set(ws, f"A{fila}", it.numero, _F_NORM, _CENTER)
        _set(ws, f"B{fila}", it.descripcion, _F_NORM, _LEFT)
        _set(ws, f"C{fila}", it.unidad, _F_NORM, _CENTER)
        _set(ws, f"D{fila}", it.cantidad, _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        _set(ws, f"E{fila}", _money(pu), _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        # Precio Total = Cantidad * Precio Unitario (fórmula viva)
        _set(ws, f"F{fila}", f"=D{fila}*E{fila}", _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        _borde_rango(ws, fila, "ABCDEF")
        filas_total.append(fila)
        fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    _set(ws, f"A{fila}", f"TOTAL GENERAL ({s})", _F_BOLD, _RIGHT, _FILL_GRIS)
    if filas_total:
        suma = "+".join(f"F{f}" for f in filas_total)
        _set(ws, f"F{fila}", f"={suma}", _F_BOLD, _RIGHT, _FILL_GRIS, _FMT_MONTO)
    else:
        _set(ws, f"F{fila}", 0, _F_BOLD, _RIGHT, _FILL_GRIS, _FMT_MONTO)
    _borde_rango(ws, fila, "ABCDEF")
    fila += 2
    _set(ws, f"A{fila}", f"Son: {_a_letras(total_general)} bolivianos.", _F_NORM,
         _LEFT_NW)
    _autoalto_filas(ws, ini_texto, fila, "B", 50)
    _config_pagina(ws, 6)
    _pie_firma(ws, fila + 2, proyecto, 6)
    return total_general


# ===================================================================== B-2
def _b2_una_hoja_por_item(wb, proyecto, items):
    """Crea UNA hoja (= una página) por cada ítem con su Formulario B-2.

    Todos los cálculos se escriben como FÓRMULAS de Excel, de modo que si el
    usuario cambia un precio o una cantidad, el APU se recalcula solo.
    """
    n = 0
    for it in items:
        recursos = repositories.listar_recursos(it.id)
        if not recursos:
            continue
        n += 1
        nombre_hoja = f"B-2 ({it.numero or n})"[:31]
        ws = wb.create_sheet(nombre_hoja)
        _b2_item(ws, proyecto, it, recursos)
    if n == 0:  # sin ítems con recursos: hoja informativa
        ws = wb.create_sheet("B-2 APU")
        fila = _encabezado(ws, "B-2", "Análisis de Precio Unitario", proyecto, 5)
        _set(ws, f"A{fila}", "Genere los APU antes de exportar.", _F_NORM, _LEFT_NW)
        _config_pagina(ws, 5)


def _b2_item(ws, proyecto, it, recursos):
    """Escribe el Formulario B-2 de un único ítem en su propia hoja, con fórmulas."""
    headers_cols = "ABCDE"
    fila = _encabezado(ws, "B-2", "Análisis de Precio Unitario", proyecto, 5)

    # Datos del ítem
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    _set(ws, f"A{fila}", f"ÍTEM {it.numero}: {it.descripcion}", _F_BOLD, _LEFT,
         _FILL_GRIS)
    _borde_rango(ws, fila, headers_cols)
    fila_desc = fila
    fila += 1
    _set(ws, f"A{fila}", "Unidad:", _F_BOLD, _LEFT_NW)
    _set(ws, f"B{fila}", it.unidad, _F_NORM, _CENTER_NW)
    _set(ws, f"C{fila}", "Cantidad:", _F_BOLD, _RIGHT_NW)
    _set(ws, f"D{fila}", it.cantidad, _F_NORM, _RIGHT_NW, fmt=_FMT_MONTO)
    fila += 2

    # Anchos de columna
    for col, w in zip("ABCDE", [46, 10, 13, 15, 16]):
        ws.column_dimensions[col].width = w

    _fila_headers(ws, fila, ["DESCRIPCION", "UNIDAD", "CANTIDAD",
                             f"PRECIO PRODUCTIVO ({_sim()})",
                             f"COSTO TOTAL ({_sim()})"],
                  [46, 10, 13, 15, 16])
    fila += 1
    ini_recursos = fila

    # parámetros del proyecto en celdas (para fórmulas legibles)
    grupos = {}
    for etiqueta, tipo in (("1. MATERIALES", TIPO_MATERIAL),
                           ("2. MANO DE OBRA", TIPO_MANO_OBRA),
                           ("3. EQUIPO, MAQUINARIA Y HERRAMIENTAS", TIPO_EQUIPO)):
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        _set(ws, f"A{fila}", etiqueta, _F_BOLD, _LEFT, _FILL_GRIS)
        _borde_rango(ws, fila, headers_cols)
        fila += 1
        filas_grupo = []
        for r in [x for x in recursos if x.tipo == tipo]:
            _set(ws, f"A{fila}", r.descripcion, _F_NORM, _LEFT)
            _set(ws, f"B{fila}", r.unidad, _F_NORM, _CENTER)
            _set(ws, f"C{fila}", r.cantidad_apu, _F_NORM, _RIGHT, fmt=_FMT_CANT)
            _set(ws, f"D{fila}", _money(r.precio_unitario), _F_NORM, _RIGHT,
                 fmt=_FMT_MONTO)
            # Parcial = Cantidad * Precio Unitario (fórmula viva)
            _set(ws, f"E{fila}", f"=C{fila}*D{fila}", _F_NORM, _RIGHT, fmt=_FMT_MONTO)
            _borde_rango(ws, fila, headers_cols)
            filas_grupo.append(fila)
            fila += 1
        # subtotal del grupo
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        etiqueta_sub = {TIPO_MATERIAL: "SUBTOTAL MATERIALES (A)",
                        TIPO_MANO_OBRA: "Subtotal mano de obra",
                        TIPO_EQUIPO: "Subtotal equipo"}[tipo]
        _set(ws, f"A{fila}", etiqueta_sub, _F_BOLD, _RIGHT)
        if filas_grupo:
            ref = f"=SUM(E{filas_grupo[0]}:E{filas_grupo[-1]})"
        else:
            ref = 0
        _set(ws, f"E{fila}", ref, _F_BOLD, _RIGHT, fmt=_FMT_MONTO)
        _borde_rango(ws, fila, headers_cols)
        grupos[tipo] = {"subtotal_row": fila}
        fila += 1

    # ---- Cargas sobre mano de obra (beneficios + IVA) ----
    r_mo_sub = grupos[TIPO_MANO_OBRA]["subtotal_row"]
    f_bs = proyecto.factor_beneficios_sociales
    f_iva = proyecto.factor_iva_mano_obra
    fila = _linea_formula(ws, fila, f"Beneficios sociales ({f_bs*100:.2f}%)",
                          f"=E{r_mo_sub}*{f_bs}")
    r_bs = fila - 1
    fila = _linea_formula(ws, fila, f"IVA mano de obra ({f_iva*100:.2f}%)",
                          f"=(E{r_mo_sub}+E{r_bs})*{f_iva}")
    r_iva = fila - 1
    fila = _linea_formula(ws, fila, "SUBTOTAL MANO DE OBRA (B)",
                          f"=E{r_mo_sub}+E{r_bs}+E{r_iva}", negrita=True, fill=True)
    r_b = fila - 1

    # ---- Cargas sobre equipo (herramientas s/ B + IVA equipo) ----
    r_eq_sub = grupos[TIPO_EQUIPO]["subtotal_row"]
    f_herr = proyecto.factor_herramientas
    f_iva_eq = proyecto.factor_iva_equipo
    fila = _linea_formula(ws, fila,
                          f"Herramientas menores ({f_herr*100:.2f}% de B)",
                          f"=E{r_b}*{f_herr}")
    r_herr = fila - 1
    extra_eq = f"+E{r_eq_sub}*{f_iva_eq}" if f_iva_eq else ""
    if f_iva_eq:
        fila = _linea_formula(ws, fila, f"IVA equipo ({f_iva_eq*100:.2f}%)",
                              f"=E{r_eq_sub}*{f_iva_eq}")
        r_iva_eq = fila - 1
        extra_eq = f"+E{r_iva_eq}"
    fila = _linea_formula(ws, fila, "SUBTOTAL EQUIPO Y HERRAMIENTAS (C)",
                          f"=E{r_eq_sub}+E{r_herr}{extra_eq}", negrita=True,
                          fill=True)
    r_c = fila - 1

    # ---- Totales finales (numeracion oficial del Formulario B-2) ----
    r_a = grupos[TIPO_MATERIAL]["subtotal_row"]
    fila = _linea_formula(ws, fila, "COSTO DIRECTO (1 + 2 + 3)",
                          f"=E{r_a}+E{r_b}+E{r_c}", negrita=True, fill=True)
    r_cd = fila - 1
    f_gg = proyecto.factor_gastos_generales
    fila = _linea_formula(ws, fila,
                          f"4. GASTOS GENERALES = {f_gg*100:.2f}% DE 1 + 2 + 3",
                          f"=E{r_cd}*{f_gg}")
    r_gg = fila - 1
    f_ut = proyecto.factor_utilidad_sabs
    fila = _linea_formula(ws, fila,
                          f"5. UTILIDAD = {f_ut*100:.2f}% DE 1 + 2 + 3 + 4",
                          f"=(E{r_cd}+E{r_gg})*{f_ut}")
    r_ut = fila - 1
    f_it = proyecto.factor_it
    fila = _linea_formula(ws, fila,
                          f"6. IMPUESTOS IT = {f_it*100:.2f}% DE 1 + 2 + 3 + 4 + 5",
                          f"=(E{r_cd}+E{r_gg}+E{r_ut})*{f_it}")
    r_it = fila - 1
    fila = _linea_formula(ws, fila,
                          f"TOTAL PRECIO UNITARIO (1+2+3+4+5+6) ({_sim()})",
                          f"=E{r_cd}+E{r_gg}+E{r_ut}+E{r_it}", negrita=True,
                          fill=True)

    _autoalto_filas(ws, ini_recursos, fila, "A", 46)
    _config_pagina(ws, 5)
    _pie_firma(ws, fila + 2, proyecto, 5)


def _linea_formula(ws, fila, etiqueta, formula, negrita=False, fill=False):
    """Escribe una línea etiqueta (A:D) + valor/fórmula (E)."""
    font = _F_BOLD if negrita else _F_NORM
    relleno = _FILL_GRIS if fill else None
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    _set(ws, f"A{fila}", etiqueta, font, _RIGHT, relleno)
    _set(ws, f"E{fila}", formula, font, _RIGHT, relleno, _FMT_MONTO)
    _borde_rango(ws, fila, "ABCDE")
    return fila + 1


# ===================================================================== B-3
def _b3_elementales(ws, proyecto, items):
    fila = _encabezado(ws, "B-3", "Precios Unitarios de Elementales", proyecto, 5)
    _fila_headers(ws, fila, ["Tipo", "Descripción del insumo", "Unidad",
                             f"Precio Unitario ({_sim()})", "Fuente del precio"],
                  [14, 46, 10, 16, 20])
    fila += 1
    vistos: dict[tuple, dict] = {}
    for it in items:
        for r in repositories.listar_recursos(it.id):
            clave = (r.tipo, r.descripcion.strip().lower(), r.unidad.strip().lower())
            if clave not in vistos or r.precio_unitario > 0:
                vistos[clave] = {"tipo": r.tipo, "desc": r.descripcion,
                                 "unidad": r.unidad, "precio": r.precio_unitario,
                                 "fuente": r.fuente_precio}
    etiqueta_tipo = {TIPO_MATERIAL: "Material", TIPO_MANO_OBRA: "Mano de obra",
                     TIPO_EQUIPO: "Equipo/Herr."}
    ini = fila
    for v in sorted(vistos.values(), key=lambda x: (x["tipo"], x["desc"])):
        _set(ws, f"A{fila}", etiqueta_tipo.get(v["tipo"], v["tipo"]), _F_NORM,
             _CENTER)
        _set(ws, f"B{fila}", v["desc"], _F_NORM, _LEFT)
        _set(ws, f"C{fila}", v["unidad"], _F_NORM, _CENTER)
        _set(ws, f"D{fila}", _money(v["precio"]), _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        _set(ws, f"E{fila}", v["fuente"] or "—", _F_NORM, _CENTER)
        _borde_rango(ws, fila, "ABCDE")
        fila += 1
    _autoalto_filas(ws, ini, fila, "B", 46)
    _config_pagina(ws, 5)
    _pie_firma(ws, fila + 2, proyecto, 5)


# ===================================================================== B-4
def _b4_equipo(ws, proyecto, items):
    fila = _encabezado(ws, "B-4", "Equipo Mínimo Comprometido", proyecto, 5)
    _fila_headers(ws, fila, ["N°", "Equipo / Maquinaria", "Unidad",
                             "Cantidad (ref.)", "Observaciones"],
                  [6, 46, 12, 14, 28])
    fila += 1
    # consolida equipos del proyecto
    equipos: dict[str, dict] = {}
    for it in items:
        for r in repositories.listar_recursos(it.id):
            if r.tipo == TIPO_EQUIPO:
                k = r.descripcion.strip().lower()
                equipos.setdefault(k, {"desc": r.descripcion, "unidad": r.unidad,
                                       "cant": 0.0})
                equipos[k]["cant"] += r.cantidad_apu * _cantidad_item(it)
    if not equipos:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
        _set(ws, f"A{fila}", "Sin equipo registrado en los APU. Complete según el "
                             "requerimiento del DBC.", _F_NORM, _LEFT_NW)
        _config_pagina(ws, 5)
        _pie_firma(ws, fila + 2, proyecto, 5)
        return
    ini = fila
    for i, v in enumerate(sorted(equipos.values(), key=lambda x: x["desc"]), 1):
        _set(ws, f"A{fila}", i, _F_NORM, _CENTER)
        _set(ws, f"B{fila}", v["desc"], _F_NORM, _LEFT)
        _set(ws, f"C{fila}", v["unidad"], _F_NORM, _CENTER)
        _set(ws, f"D{fila}", round(v["cant"], 2), _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        _set(ws, f"E{fila}", "Propio / Alquilado", _F_NORM, _LEFT)
        _borde_rango(ws, fila, "ABCDE")
        fila += 1
    _autoalto_filas(ws, ini, fila, "B", 46)
    _config_pagina(ws, 5)
    _pie_firma(ws, fila + 2, proyecto, 5)


# ============================================================ cronograma base
def _calcular_cronograma(proyecto, items):
    """Calcula el cronograma con SECUENCIA LÓGICA de obra (CPM/Gantt).

    Usa core.cronograma para clasificar cada ítem en su fase constructiva
    (replanteo → preparación → estructura de abajo hacia arriba → instalaciones
    → acabados) y programarlo dentro del plazo contractual.

    Devuelve (n_periodos, datos) donde cada dato tiene:
      {item, mes_inicio, mes_fin, fase, monto_total}
    """
    from core.cronograma import construir_cronograma

    plazo = int(getattr(proyecto, "plazo_dias", 180) or 180)
    montos = {}
    for it in items:
        if not it.descripcion:
            continue
        res = repositories.obtener_resultado(it.id)
        pu = res.precio_unitario_total if res else 0.0
        montos[it.id] = pu * (it.cantidad or 0.0)

    n_periodos, programadas = construir_cronograma(items, plazo, montos)
    datos = []
    for a in programadas:
        datos.append({"item": a.item, "mes_inicio": a.mes_inicio,
                      "mes_fin": a.mes_fin, "fase": a.fase,
                      "monto_total": a.monto})
    return n_periodos, datos


def _montos_por_periodo(n_periodos, datos):
    """Distribuye el monto de cada actividad entre sus meses de ejecución."""
    montos = [0.0] * (n_periodos + 1)  # índice 0 sin uso
    for d in datos:
        m_ini = max(1, d.get("mes_inicio", 1))
        m_fin = max(m_ini, min(d.get("mes_fin", m_ini), n_periodos))
        span = m_fin - m_ini + 1
        cuota = d["monto_total"] / span if span else d["monto_total"]
        for m in range(m_ini, m_fin + 1):
            montos[m] += cuota
    return montos


# ===================================================================== A-8
def _a8_cronograma_obra(ws, proyecto, items):
    n_periodos, datos = _calcular_cronograma(proyecto, items)
    ncols = 5 + n_periodos
    fila = _encabezado(ws, "A-8", "Cronograma de Ejecución de Obra", proyecto, ncols)
    _set(ws, f"A{fila-1}", f"Plazo contractual: {proyecto.plazo_dias} días "
                           f"calendario ({n_periodos} meses). Secuencia lógica "
                           f"(replanteo → preparación → estructura → instalaciones "
                           f"→ acabados), barras tipo Gantt (CPM).", _F_NORM, _LEFT_NW)
    fila += 1
    headers = ["N°", "Ítem / Actividad", "Fase", "Unidad", "Cantidad"] + \
              [f"Mes {i}" for i in range(1, n_periodos + 1)]
    anchos = [6, 34, 18, 8, 10] + [7] * n_periodos
    _fila_headers(ws, fila, headers, anchos)
    fila += 1
    ini_texto = fila
    fill_barra = PatternFill("solid", fgColor="9DC3E6")
    for d in datos:
        it = d["item"]
        _set(ws, f"A{fila}", it.numero, _F_NORM, _CENTER)
        _set(ws, f"B{fila}", it.descripcion, _F_NORM, _LEFT)
        _set(ws, f"C{fila}", d["fase"], _F_NORM, _LEFT)
        _set(ws, f"D{fila}", it.unidad, _F_NORM, _CENTER)
        _set(ws, f"E{fila}", it.cantidad, _F_NORM, _RIGHT, fmt=_FMT_MONTO)
        m_ini, m_fin = d["mes_inicio"], d["mes_fin"]
        for m in range(1, n_periodos + 1):
            col = get_column_letter(5 + m)
            en_barra = m_ini <= m <= m_fin
            c = _set(ws, f"{col}{fila}", "■" if en_barra else "", _F_NORM, _CENTER)
            if en_barra:
                c.fill = fill_barra
            c.border = _BORDE
        for c in "ABCDE":
            ws[f"{c}{fila}"].border = _BORDE
        fila += 1

    # Fila de avance valorado y % por período + acumulado
    montos = _montos_por_periodo(n_periodos, datos)
    total = sum(montos)
    _set(ws, f"A{fila}", f"AVANCE VALORADO ({_sim()})", _F_BOLD, _RIGHT, _FILL_GRIS)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    for m in range(1, n_periodos + 1):
        col = get_column_letter(5 + m)
        _set(ws, f"{col}{fila}", _money(montos[m]), _F_BOLD, _RIGHT, _FILL_GRIS,
             "#,##0")
    fila += 1
    _set(ws, f"A{fila}", "AVANCE PARCIAL (%)", _F_BOLD, _RIGHT, _FILL_GRIS)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    acum = 0.0
    fila_acum = fila + 1
    for m in range(1, n_periodos + 1):
        col = get_column_letter(5 + m)
        pct = (montos[m] / total * 100) if total else 0.0
        _set(ws, f"{col}{fila}", round(pct, 1), _F_BOLD, _RIGHT, _FILL_GRIS, "0.0")
    fila += 1
    _set(ws, f"A{fila}", "AVANCE ACUMULADO (%)", _F_BOLD, _RIGHT, _FILL_GRIS)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    for m in range(1, n_periodos + 1):
        col = get_column_letter(5 + m)
        acum += (montos[m] / total * 100) if total else 0.0
        _set(ws, f"{col}{fila}", round(acum, 1), _F_BOLD, _RIGHT, _FILL_GRIS, "0.0")
    fila += 2
    _autoalto_filas(ws, ini_texto, ini_texto + len(datos), "B", 34)
    _config_pagina(ws, ncols)
    _pie_firma(ws, fila, proyecto, ncols)


# ===================================================================== A-9
def _a9_movilizacion(ws, proyecto, items):
    """Cronograma de movilización de equipos (en función del equipo del B-2)."""
    n_periodos, datos = _calcular_cronograma(proyecto, items)
    ncols = 2 + n_periodos
    fila = _encabezado(ws, "A-9", "Cronograma de Movilización de Equipo",
                       proyecto, ncols)
    fila += 1
    headers = ["N°", "Equipo / Maquinaria"] + [f"Mes {i}" for i in
                                               range(1, n_periodos + 1)]
    anchos = [6, 40] + [8] * n_periodos
    _fila_headers(ws, fila, headers, anchos)
    fila += 1

    # Para cada equipo (del B-2), determinar en qué meses se usa según el A-8:
    # el rango de meses de TODAS las actividades que usan ese equipo.
    meses_por_item = {d["item"].id: (d["mes_inicio"], d["mes_fin"]) for d in datos}
    equipos: dict[str, set] = {}
    for it in items:
        for r in repositories.listar_recursos(it.id):
            if r.tipo == TIPO_EQUIPO:
                k = r.descripcion.strip()
                equipos.setdefault(k, set())
                if it.id in meses_por_item:
                    mi, mf = meses_por_item[it.id]
                    equipos[k].update(range(mi, mf + 1))

    if not equipos:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncols)
        _set(ws, f"A{fila}", "Sin equipo registrado en los APU (Formulario B-2).",
             _F_NORM, _LEFT_NW)
        fila += 2
        _config_pagina(ws, ncols)
        _pie_firma(ws, fila, proyecto, ncols)
        return

    fill_barra = PatternFill("solid", fgColor="C6E0B4")
    for i, (nombre, meses) in enumerate(sorted(equipos.items()), 1):
        _set(ws, f"A{fila}", i, _F_NORM, _CENTER)
        _set(ws, f"B{fila}", nombre, _F_NORM, _LEFT)
        # el equipo permanece movilizado desde el primer hasta el último mes de uso
        m_ini, m_fin = (min(meses), max(meses)) if meses else (1, 1)
        for m in range(1, n_periodos + 1):
            col = get_column_letter(2 + m)
            en_barra = m_ini <= m <= m_fin
            c = _set(ws, f"{col}{fila}", "■" if en_barra else "", _F_NORM, _CENTER)
            if en_barra:
                c.fill = fill_barra
            c.border = _BORDE
        for c in ("A", "B"):
            ws[f"{c}{fila}"].border = _BORDE
        fila += 1
    fila += 1
    _set(ws, f"A{fila}", "Nota: la movilización abarca desde el primer hasta el "
                         "último mes de uso de cada equipo según el A-8.", _F_NORM,
         _LEFT_NW)
    fila += 2
    _config_pagina(ws, ncols)
    _pie_firma(ws, fila, proyecto, ncols)


# ===================================================================== B-5
def _b5_desembolsos(ws, proyecto, items):
    """Cronograma de desembolsos. Depende del A-8 y del anticipo solicitado."""
    n_periodos, datos = _calcular_cronograma(proyecto, items)
    fila = _encabezado(ws, "B-5", "Cronograma de Desembolsos", proyecto, 5)

    montos = _montos_por_periodo(n_periodos, datos)
    total_obra = sum(montos)

    anticipo = (total_obra * proyecto.porcentaje_anticipo
                if proyecto.solicita_anticipo else 0.0)

    # Mensaje de anticipo
    s = _sim()
    if proyecto.solicita_anticipo:
        _set(ws, f"A{fila}", f"Anticipo solicitado: SÍ  ·  "
                             f"{proyecto.porcentaje_anticipo*100:.0f}% = "
                             f"{s} {_money(anticipo):,.2f}", _F_BOLD, _LEFT_NW)
    else:
        _set(ws, f"A{fila}", "Anticipo solicitado: NO", _F_BOLD, _LEFT_NW)
    fila += 2

    _fila_headers(ws, fila, ["Período", f"Avance valorado ({s})",
                             f"Amortización anticipo ({s})",
                             f"Desembolso neto ({s})", f"Acumulado ({s})"],
                  [22, 20, 22, 20, 20])
    fila += 1

    # Fila de anticipo (al inicio, si aplica)
    acumulado = 0.0
    if anticipo:
        acumulado += anticipo
        _set(ws, f"A{fila}", "Anticipo (inicio)", _F_BOLD, _LEFT)
        _set(ws, f"B{fila}", 0.0, _F_NORM, _RIGHT, fmt="#,##0.00")
        _set(ws, f"C{fila}", 0.0, _F_NORM, _RIGHT, fmt="#,##0.00")
        _set(ws, f"D{fila}", _money(anticipo), _F_BOLD, _RIGHT, fmt="#,##0.00")
        _set(ws, f"E{fila}", _money(acumulado), _F_NORM, _RIGHT, fmt="#,##0.00")
        for c in "ABCDE":
            ws[f"{c}{fila}"].border = _BORDE
        fila += 1

    # Desembolsos por mes: avance del mes menos amortización proporcional del anticipo
    for m in range(1, n_periodos + 1):
        avance = montos[m]
        # amortización proporcional al avance del período
        amort = (anticipo * (avance / total_obra)) if total_obra else 0.0
        desembolso = avance - amort
        acumulado += desembolso
        _set(ws, f"A{fila}", f"Mes {m}", _F_NORM, _LEFT)
        _set(ws, f"B{fila}", _money(avance), _F_NORM, _RIGHT, fmt="#,##0.00")
        _set(ws, f"C{fila}", _money(amort), _F_NORM, _RIGHT, fmt="#,##0.00")
        _set(ws, f"D{fila}", _money(desembolso), _F_NORM, _RIGHT, fmt="#,##0.00")
        _set(ws, f"E{fila}", _money(acumulado), _F_NORM, _RIGHT, fmt="#,##0.00")
        for c in "ABCDE":
            ws[f"{c}{fila}"].border = _BORDE
        fila += 1

    # Totales
    _set(ws, f"A{fila}", "TOTAL", _F_BOLD, _RIGHT, _FILL_GRIS)
    _set(ws, f"B{fila}", _money(total_obra), _F_BOLD, _RIGHT, _FILL_GRIS, "#,##0.00")
    _set(ws, f"C{fila}", _money(anticipo), _F_BOLD, _RIGHT, _FILL_GRIS, "#,##0.00")
    _set(ws, f"D{fila}", _money(total_obra - anticipo), _F_BOLD, _RIGHT,
         _FILL_GRIS, "#,##0.00")
    _set(ws, f"E{fila}", _money(acumulado), _F_BOLD, _RIGHT, _FILL_GRIS, "#,##0.00")
    for c in "ABCDE":
        ws[f"{c}{fila}"].border = _BORDE
    fila += 2
    _config_pagina(ws, 5)
    _pie_firma(ws, fila, proyecto, 5)


# ---------------------------------------------------------------- pie de firma
def _pie_firma(ws, fila, proyecto, ncols):
    """Bloque de firma del representante legal al pie del formulario."""
    fila += 2
    centro = max(2, ncols // 2)
    col = get_column_letter(centro)
    _set(ws, f"{col}{fila}", "_______________________________", _F_NORM, _CENTER_NW)
    fila += 1
    nombre = proyecto.representante_legal or "(Representante Legal)"
    _set(ws, f"{col}{fila}", nombre, _F_BOLD, _CENTER_NW)
    fila += 1
    if proyecto.ci_representante:
        _set(ws, f"{col}{fila}", f"C.I. {proyecto.ci_representante}", _F_NORM,
             _CENTER_NW)
        fila += 1
    _set(ws, f"{col}{fila}", "REPRESENTANTE LEGAL", _F_NORM, _CENTER_NW)
    fila += 1
    empresa = proyecto.proponente or proyecto.nombre
    _set(ws, f"{col}{fila}", empresa, _F_NORM, _CENTER_NW)


# ---------------------------------------------------------------- helpers
def _modulo_nombre(item):
    """Devuelve el nombre del módulo del ítem o None."""
    if not item.modulo_id:
        return None
    try:
        from core.database import db_session
        with db_session() as conn:
            r = conn.execute("SELECT nombre FROM modulos WHERE id=?",
                             (item.modulo_id,)).fetchone()
            return r["nombre"] if r else None
    except Exception:
        return None


def _cantidad_item(item):
    return item.cantidad or 0.0


def _a_letras(monto: float) -> str:
    """Convierte un monto a una expresión simple (entero + centavos)."""
    entero = int(monto)
    centavos = int(round((monto - entero) * 100))
    return f"{entero:,} con {centavos:02d}/100"


# ---------------------------------------------------------------- API
def exportar_formularios(proyecto_id: int, ruta: str | Path | None = None) -> Path:
    """Genera el libro Excel con los Formularios oficiales NB-SABS (DS 0181):

      B-1 Presupuesto · B-2 APU · B-3 Elementales · B-4 Equipo Mínimo
      A-8 Cronograma de Obra · A-9 Movilización de Equipo · B-5 Desembolsos
    """
    proyecto = repositories.obtener_proyecto(proyecto_id)
    if not proyecto:
        raise ValueError(f"Proyecto {proyecto_id} no encontrado")
    if ruta is None:
        ruta = settings.EXPORT_DIR / f"formularios_proyecto_{proyecto_id}.xlsx"
    ruta = Path(ruta)

    items = repositories.listar_items(proyecto_id)

    # Fijar la moneda del libro: los montos internos están en BOB; si el
    # proyecto usa USD, se convierten dividiendo por el tipo de cambio.
    from core import currency
    if proyecto.moneda == currency.USD:
        tc = currency.tipo_cambio(proyecto) or 6.96
        _MONEDA["sim"], _MONEDA["factor"] = "$us", 1.0 / tc
    else:
        _MONEDA["sim"], _MONEDA["factor"] = "Bs", 1.0

    wb = Workbook()
    wb.remove(wb.active)
    # Precisión de pantalla: Excel calcula con los valores tal como se muestran
    # (fullPrecision=False) y recalcula las fórmulas al abrir el archivo.
    wb.calculation.fullPrecision = False
    wb.calculation.fullCalcOnLoad = True

    _b1_presupuesto(wb.create_sheet("B-1 Presupuesto"), proyecto, items)
    _b2_una_hoja_por_item(wb, proyecto, items)   # un B-2 por ítem (una pág. c/u)
    _b3_elementales(wb.create_sheet("B-3 Elementales"), proyecto, items)
    _b4_equipo(wb.create_sheet("B-4 Equipo"), proyecto, items)
    _a8_cronograma_obra(wb.create_sheet("A-8 Cronograma Obra"), proyecto, items)
    _a9_movilizacion(wb.create_sheet("A-9 Movilizacion"), proyecto, items)
    _b5_desembolsos(wb.create_sheet("B-5 Desembolsos"), proyecto, items)
    wb.save(ruta)
    return ruta
