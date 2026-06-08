"""Prueba del importador de B-2 en formato vertical (ACTIVIDAD:/UNITARIO:...)."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from scripts.importar_apu_banco import importar  # noqa: E402


def _crear_xlsx(ruta):
    wb = Workbook()
    wb.active.title = "PRESUPUESTO"  # primera hoja (resumen, sin APU)
    ws = wb.create_sheet("APU")
    filas = [
        [None, "ANÁLISIS DE PRECIO UNITARIO"],
        ["ACTIVIDAD:     1.1    Hormigón Armado"],
        ["UNITARIO:     M3"],
        ["CANTIDAD:     25"],
        ["Moneda: Bolivianos"],
        [None, "Descripción", "Und.", "Cantidad", "%", "P.Improd", "P.Prod", "Total"],
        ["1.- ", "MATERIALES"],
        [None, "Cemento Portland", "kg", 320, None, None, 1.2, 384],
        [None, None, None, None, None, None, "TOTAL MATERIALES", 384],
        ["2.- ", "MANO DE OBRA"],
        [None, "Albañil", "hr", 8, None, None, 18.75, 150],
        [None, "BENEFICIOS SOCIALES - % ", None, None, None, 0.55, 82.5, None],
        [None, None, None, None, None, None, "SUBTOTAL MANO DE OBRA", 150],
        ["3.- ", "EQUIPO Y HERRAMIENTAS"],
        [None, "Mezcladora", "hr", 2, None, None, 25, 50],
        [None, "HERRAMIENTAS - % DEL TOTAL", None, None, None, 0.05, 7.5, None],
        ["4.- ", "GASTOS GENERALES"],
        [None, "GASTOS GENERALES - % DE 1+2+3", None, None, None, 0.1, 58, None],
    ]
    for f in filas:
        ws.append(f)
    wb.save(ruta)


def test_importa_formato_vertical(tmp_path):
    ruta = tmp_path / "vertical.xlsx"
    _crear_xlsx(str(ruta))
    apus = importar(str(ruta))
    assert len(apus) == 1
    a = apus[0]
    assert "Hormigón Armado" in a["actividad"]
    assert a["unidad"] == "M3"
    assert a["cantidad"] == 25
    # recursos reales, sin colar totales ni lineas de %
    assert [m["descripcion"] for m in a["materiales"]] == ["Cemento Portland"]
    assert [m["descripcion"] for m in a["mano_obra"]] == ["Albañil"]
    assert [m["descripcion"] for m in a["equipo"]] == ["Mezcladora"]
    assert a["materiales"][0]["precio"] == 1.2
    assert a["materiales"][0]["cantidad"] == 320
    assert a["mano_obra"][0]["unidad"] == "hr"


def _crear_xlsx_flexible(ruta):
    """Formato B-2 oficial desplazado: col A vacia, etiquetas 'Actividad :' en B,
    numero de item antes de la descripcion."""
    wb = Workbook()
    ws = wb.active
    ws.title = "B-2"
    filas = [
        [None, "FORMULARIO B-2"],
        [None, "ANÁLISIS DE PRECIOS UNITARIOS"],
        [None, "Proyecto :", "OBRA X"],
        [None, "Actividad :", "1. Provisión e instalación de cable"],
        [None, "Cantidad  :", 53],
        [None, "Unidad :", "PIEZA"],
        [None, "Moneda :", "BOLIVIANOS"],
        [None, "1.        MATERIALES"],
        [None, "DESCRIPCIÓN", None, "UNIDAD", "CANTIDAD", "PRECIO PRODUCTIVO",
         "COSTO TOTAL"],
        [None, "1", "Cable UTP Cat 6", "Mts", 45, 4.74, 213.3],
        [None, "2", "Conector Jack RJ-45", "Pza", 1, 38.4, 38.4],
        [None, "TOTAL MATERIALES", None, None, None, None, 251.7],
        [None, "2.        MANO DE OBRA"],
        [None, "DESCRIPCIÓN", None, "UNIDAD", "CANTIDAD", "PRECIO PRODUCTIVO",
         "COSTO TOTAL"],
        [None, "1", "Tecnico Especialista", "HR", 1.88, 30, 56.6],
        [None, "CARGAS SOCIALES = (% ", None, None, 0.7118, 40.3],
        [None, "TOTAL MANO DE OBRA", None, None, None, None, 96.9],
        [None, "3.        EQUIPO, MAQUINARIA Y HERRAMIENTAS"],
        [None, "DESCRIPCIÓN", None, "UNIDAD", "CANTIDAD", "PRECIO PRODUCTIVO",
         "COSTO TOTAL"],
        [None, "0.03", "HERRAMIENTAS = (% DE M.O.)", None, None, 2.9],
        [None, "4.        GASTOS GENERALES"],
        [None, "0.07", "GASTOS GENERALES = % DE 1+2+3", None, None, 31.9],
    ]
    for f in filas:
        ws.append(f)
    wb.save(ruta)


def test_importa_formato_flexible(tmp_path):
    ruta = tmp_path / "flexible.xlsx"
    _crear_xlsx_flexible(str(ruta))
    apus = importar(str(ruta))
    assert len(apus) == 1
    a = apus[0]
    assert a["actividad"].startswith("1. Provisión")
    assert a["unidad"] == "PIEZA"
    assert a["cantidad"] == 53
    # toma la descripcion (no el numero de item) y excluye totales/% /herramientas
    assert [m["descripcion"] for m in a["materiales"]] == [
        "Cable UTP Cat 6", "Conector Jack RJ-45"]
    assert [m["descripcion"] for m in a["mano_obra"]] == ["Tecnico Especialista"]
    assert a["equipo"] == []
    assert a["materiales"][0]["unidad"] == "Mts"
    assert a["materiales"][0]["cantidad"] == 45
    assert a["materiales"][0]["precio"] == 4.74


def _crear_xlsx_europeo(ruta):
    """Formato con etiqueta 'Item:', secciones por letra (A/B/C) y numeros en
    formato europeo (1.234,56) guardados como texto."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    filas = [
        [None, "Item: Cimiento de Hº Cº", None, "Unidad: m³"],
        [None, "Proyecto: Obra Y"],
        [None, "Módulo: (M01) - Muros"],
        ["Nº", "P.", "Insumo/Parámetro", "Und.", "Cant.", "Unit. (Bs)",
         "Parcial (Bs)"],
        [">", "A", "MATERIALES", None, None, None, None],
        ["1", "-", "CEMENTO PORTLAND", "kg", "120,00", "1.100,00", "132.000,00"],
        ["2", "-", "ARENA COMUN", "m³", "0,20", "120,00", "24,00"],
        [">", "D", "TOTAL MATERIALES", None, None, "(A) =", "288,00"],
        [None, "B", "MANO DE OBRA", None, None, None, None],
        ["1", "-", "ALBAÑIL", "hr", "5,00", "20,00", "100,00"],
        [">", "E", "SUBTOTAL MANO DE OBRA", None, None, "(B) =", "175,00"],
        [None, "F", "Beneficios Sociales", None, "55,00% de", "(B) =", "96,25"],
        [None, "C", "EQUIPO, MAQUINARIA Y HERRAMIENTAS", None, None, None, None],
        ["1", "-", "MEZCLADORA", "hr", "2,00", "1.250,00", "2.500,00"],
        [None, "H", "Herramientas menores", None, "5,00% de", "(G) =", "8,75"],
    ]
    for f in filas:
        ws.append(f)
    wb.save(ruta)


def test_importa_formato_europeo_por_letras(tmp_path):
    ruta = tmp_path / "europeo.xlsx"
    _crear_xlsx_europeo(str(ruta))
    apus = importar(str(ruta))
    assert len(apus) == 1
    a = apus[0]
    assert a["actividad"].startswith("Cimiento")
    assert a["unidad"] == "m³"
    # numeros europeos parseados (1.100,00 -> 1100.0; 0,20 -> 0.2)
    assert [m["descripcion"] for m in a["materiales"]] == [
        "CEMENTO PORTLAND", "ARENA COMUN"]
    assert a["materiales"][0]["precio"] == 1100.0
    assert a["materiales"][1]["cantidad"] == 0.2
    assert [m["descripcion"] for m in a["mano_obra"]] == ["ALBAÑIL"]
    assert [m["descripcion"] for m in a["equipo"]] == ["MEZCLADORA"]
    assert a["equipo"][0]["precio"] == 1250.0
    # nada de totales/% colado
    todas = [m["descripcion"] for g in ("materiales", "mano_obra", "equipo")
             for m in a[g]]
    assert not any("TOTAL" in d.upper() or "Beneficios" in d or "Herramientas" in d
                   for d in todas)


def _crear_xlsx_titulado(ruta):
    """Formato delimitado por el titulo, con actividad y unidad en lineas sin
    etiqueta y precio en la subcolumna 'Productivo'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis de Precio Unitario"
    filas = [
        [None, "ANALISIS DE PRECIO UNITARIO"],
        [None, "PROYECTO: PTAR San Pablo"],
        [None, "1 - INSTALACION DE FAENAS"],
        [None, "GLB"],
        [None, "Cantidad: 1.00"],
        [None, "Moneda: BS"],
        [None, "Descripción", "Und.", "Cantidad", "% Productiv.",
         "Precio Unitario", None, None, "Costo Total"],
        [None, None, None, None, None, "Improductivo", "Productiv.", None, None],
        [None, "1.- MATERIALES E INSUMOS"],
        [None, "CALAMINA GALVANIZADA", "M2", 26.5, None, None, 22.6723, "M",
         600.81],
        [None, "CLAVOS DE CALAMINA", "KG", 5.3, None, None, 15.77, "M", 83.58],
        [None, "TOTAL MATERIALES E INSUMOS", None, None, None, None, None, None,
         684.4],
        [None, "2.- MANO DE OBRA"],
        [None, "ALBANIL", "HR.", 139.295, None, None, 20.0998, "O", 2799.8],
        [None, "TOTAL MANO DE OBRA", None, None, None, None, None, None, 2799.8],
        [None, "3.- EQUIPO MAQUINARIA Y HERRAMIENTAS"],
        [None, "HERRAMIENTAS - % A LA M.O.", None, None, None, 0.05, None, None,
         140.0],
    ]
    for f in filas:
        ws.append(f)
    wb.save(ruta)


def test_importa_formato_titulado(tmp_path):
    ruta = tmp_path / "titulado.xlsx"
    _crear_xlsx_titulado(str(ruta))
    apus = importar(str(ruta))
    assert len(apus) == 1
    a = apus[0]
    assert a["actividad"] == "1 - INSTALACION DE FAENAS"
    assert a["unidad"] == "GLB"
    # precio tomado de la subcolumna Productivo, no de Improductivo (vacia)
    assert [m["descripcion"] for m in a["materiales"]] == [
        "CALAMINA GALVANIZADA", "CLAVOS DE CALAMINA"]
    assert a["materiales"][0]["precio"] == 22.6723
    assert a["materiales"][0]["cantidad"] == 26.5
    assert [m["descripcion"] for m in a["mano_obra"]] == ["ALBANIL"]
    assert a["equipo"] == []  # solo habia la linea de herramientas (%)
