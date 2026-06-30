"""Estilo del exportador: el texto FUERA de las tablas no ajusta texto (wrap)."""
import sys
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

import exporters.export_formularios as ef  # noqa: E402


def _proyecto():
    return SimpleNamespace(nombre="Escuela", entidad="GAM", moneda="BOB",
                           tipo_cambio=6.96, proponente="Constructora X",
                           region="La Paz")


def test_encabezado_sin_ajustar_texto():
    wb = Workbook()
    ws = wb.active
    ef._encabezado(ws, "B-2", "Análisis de Precio Unitario", _proyecto(), 5)
    for celda in ("A1", "A2", "A3", "A4"):
        assert ws[celda].alignment.wrap_text is False


def test_pie_firma_sin_ajustar_texto():
    wb = Workbook()
    ws = wb.active
    proy = SimpleNamespace(representante_legal="Juan Pérez", ci_representante="123",
                           proponente="Constructora X", nombre="Escuela")
    ef._pie_firma(ws, 10, proy, 5)
    celdas = [c for row in ws.iter_rows() for c in row if c.value]
    assert celdas, "el pie de firma debe escribir celdas"
    assert all(c.alignment.wrap_text is False for c in celdas)


def test_tablas_conservan_ajuste_de_texto():
    # Dentro de las tablas se mantiene el wrap (descripciones largas).
    assert ef._CENTER.wrap_text is True
    assert ef._LEFT.wrap_text is True
    assert ef._RIGHT.wrap_text is True
    # Las variantes "NW" efectivamente NO ajustan texto.
    assert ef._CENTER_NW.wrap_text is False
    assert ef._LEFT_NW.wrap_text is False
    assert ef._RIGHT_NW.wrap_text is False
