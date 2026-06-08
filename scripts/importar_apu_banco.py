"""Importa un Formulario B-2 (APU) oficial como banco de conocimiento.

Extrae cada actividad con sus materiales, mano de obra y equipo (descripcion,
codigo, unidad, cantidad y precio productivo) y los guarda en
data/banco_apu.json. Este banco sirve de:
  - referencia de RENDIMIENTOS y PRECIOS para el motor de APU,
  - ejemplos para la IA,
  - precios elementales para el cotizador (Nivel 1: base propia).

Uso:
    python -m scripts.importar_apu_banco <ruta_excel> [--proyecto "Nombre"]
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from config import settings  # noqa: E402
from core.text_cleaner import normalizar  # noqa: E402

# Mapeo de la seccion al tipo de recurso del programa.
_TIPO = {"MATERIALES": "material", "MANO DE OBRA": "mano_obra",
         "EQUIPO": "equipo"}


def _txt(v):
    return str(v).strip() if v is not None else ""


# Estilo numerico de la hoja en curso: 'us' (1,234.56) o 'eu' (1.234,56).
# Se autodetecta por hoja en importar(); _num lo respeta.
_ESTILO_NUM = "us"


def _detectar_estilo(filas: list) -> str:
    """Detecta si la hoja usa coma o punto como separador decimal.

    Mira el ULTIMO separador de cada celda numerica: si predominan las comas
    como ultimo separador, es formato europeo ('eu')."""
    import re as _re
    eu = us = 0
    for row in filas:
        for c in row:
            s = str(c) if c is not None else ""
            if not _re.search(r"\d", s):
                continue
            lc, lp = s.rfind(","), s.rfind(".")
            if lc < 0 and lp < 0:
                continue
            if lc > lp:
                eu += 1
            elif lp > lc:
                us += 1
    return "eu" if eu > us else "us"


def _num(v):
    if v is None:
        return 0.0
    # openpyxl ya devuelve numeros reales como float/int.
    if isinstance(v, (int, float)):
        return float(v)
    import re as _re
    s = _re.sub(r"[^\d,.\-]", "", str(v).strip())
    if not s or s in ("-", ".", ","):
        return 0.0
    if _ESTILO_NUM == "eu":
        s = s.replace(".", "").replace(",", ".")  # punto=miles, coma=decimal
    else:
        if "," in s and "." in s:
            s = s.replace(",", "")                 # coma=miles, punto=decimal
        else:
            s = s.replace(",", ".")                # solo coma -> decimal
    try:
        return float(s)
    except ValueError:
        return 0.0


def importar(ruta: str) -> list[dict]:
    """Devuelve la lista de APUs extraidos del Excel.

    Recorre TODAS las hojas y reconoce tres formatos de Formulario B-2:
      - 'estandar': etiquetas en celdas ('Actividad' en col C, valor en col E).
      - 'vertical': cabecera 'ACTIVIDAD:/UNITARIO:/CANTIDAD:' en col A y
        secciones '1.- MATERIALES', '2.- MANO DE OBRA', '3.- EQUIPO...'.
      - 'flexible': cabecera 'Actividad :/Unidad :/Cantidad :' y fila de
        encabezado 'DESCRIPCION | UNIDAD | CANTIDAD | PRECIO...' desde la que se
        detectan las columnas dinamicamente (sirve para variantes desplazadas).
    """
    global _ESTILO_NUM
    wb = load_workbook(ruta, data_only=True)
    apus: list[dict] = []
    for hoja in wb.sheetnames:
        filas = list(wb[hoja].iter_rows(values_only=True))
        _ESTILO_NUM = _detectar_estilo(filas)
        extraidos = _extraer_estandar(filas)
        if not extraidos:
            extraidos = _extraer_vertical(filas)
        if not extraidos:
            extraidos = _extraer_flexible(filas)
        apus.extend(extraidos)
    return apus


def _extraer_estandar(filas: list) -> list[dict]:
    """Formato con etiquetas en celdas: 'Actividad'/'Unidad'/'Cantidad' en col C."""
    apus: list[dict] = []
    actual = None
    seccion = None  # material | mano_obra | equipo
    for row in filas:
        cols = [_txt(c) for c in row]
        # texto util de la fila (saltando vacios)
        a = cols[0] if len(cols) > 0 else ""
        c = cols[2] if len(cols) > 2 else ""
        e = cols[4] if len(cols) > 4 else ""
        f = cols[5] if len(cols) > 5 else ""
        h = cols[7] if len(cols) > 7 else ""
        i = cols[8] if len(cols) > 8 else ""

        # Nueva actividad
        if c == "Actividad" and e:
            if actual and (actual["materiales"] or actual["mano_obra"]
                           or actual["equipo"]):
                apus.append(actual)
            actual = {"actividad": e, "unidad": "", "cantidad": 1.0,
                      "materiales": [], "mano_obra": [], "equipo": []}
            seccion = None
            continue
        if actual is None:
            continue
        if c == "Unidad" and e:
            actual["unidad"] = e
            continue
        if c == "Cantidad" and e:
            actual["cantidad"] = _num(e) or 1.0
            continue

        # Cambios de seccion (clave del dict del APU)
        cu = normalizar(c)
        if "materiales" in cu and a == "1":
            seccion = "materiales"
            continue
        if "mano de obra" in cu and a == "2":
            seccion = "mano_obra"
            continue
        if ("equipo" in cu or "maquinaria" in cu) and a == "3":
            seccion = "equipo"
            continue
        # Fin del APU al llegar a Gastos generales
        if "gastos generales" in cu and a in ("4", "5", "6"):
            seccion = None
            continue

        # Fila de recurso: tiene codigo en B y descripcion en C, unidad en F
        b = cols[1] if len(cols) > 1 else ""
        if seccion and b and c and f and a not in ("*", ""):
            actual[seccion].append({
                "codigo": b, "descripcion": c, "unidad": f,
                "cantidad": _num(h), "precio": _num(i)})

    if actual and (actual["materiales"] or actual["mano_obra"]
                   or actual["equipo"]):
        apus.append(actual)
    return apus


# Descripciones que NO son recursos sino lineas de calculo/totales del B-2.
_EXCLUIR_VERTICAL = (
    "total", "subtotal", "beneficios sociales", "impuesto al valor",
    "impuesto a las transacc", "gastos generales", "utilidad", "herramientas - %",
)


def _es_recurso_vertical(desc: str) -> bool:
    dn = normalizar(desc)
    if not dn:
        return False
    if "%" in desc:  # lineas de porcentaje (beneficios, IVA, herramientas, IT...)
        return False
    return not any(k in dn for k in _EXCLUIR_VERTICAL)


def _valor_etiqueta(texto: str) -> str:
    """De 'ACTIVIDAD:   1.1   Instalacion' devuelve 'Instalacion' (tras los ':')."""
    val = texto.split(":", 1)[1] if ":" in texto else texto
    return " ".join(val.split()).strip()


def _extraer_vertical(filas: list) -> list[dict]:
    """Formato vertical: 'ACTIVIDAD:/UNITARIO:/CANTIDAD:' en col A y secciones
    '1.- MATERIALES', '2.- MANO DE OBRA', '3.- EQUIPO Y HERRAMIENTAS'.

    Columnas de recurso (0-index): 1=Descripcion, 2=Unidad, 3=Cantidad,
    6=Precio Productivo (unitario)."""
    apus: list[dict] = []
    actual = None
    seccion = None
    for row in filas:
        cols = [_txt(c) for c in row]
        a = cols[0] if cols else ""
        an = normalizar(a)
        b = cols[1] if len(cols) > 1 else ""
        bn = normalizar(b)

        if an.startswith("actividad"):
            if actual and (actual["materiales"] or actual["mano_obra"]
                           or actual["equipo"]):
                apus.append(actual)
            actual = {"actividad": _valor_etiqueta(a) or "Actividad",
                      "unidad": "", "cantidad": 1.0,
                      "materiales": [], "mano_obra": [], "equipo": []}
            seccion = None
            continue
        if actual is None:
            continue
        if an.startswith("unitario") or an.startswith("unidad"):
            actual["unidad"] = _valor_etiqueta(a)
            continue
        if an.startswith("cantidad"):
            actual["cantidad"] = _num(_valor_etiqueta(a)) or 1.0
            continue

        # Cambios de seccion: 'N.-' en col A y nombre en col B.
        if a.startswith("1") and "material" in bn:
            seccion = "materiales"
            continue
        if a.startswith("2") and "mano de obra" in bn:
            seccion = "mano_obra"
            continue
        if a.startswith("3") and ("equipo" in bn or "maquinaria" in bn
                                  or "herramient" in bn):
            seccion = "equipo"
            continue
        if a.startswith(("4", "5", "6")) and b:
            seccion = None
            continue

        # Fila de recurso del formato vertical.
        if seccion and _es_recurso_vertical(b):
            unidad = cols[2] if len(cols) > 2 else ""
            cantidad = _num(cols[3]) if len(cols) > 3 else 0.0
            precio = _num(cols[6]) if len(cols) > 6 else 0.0
            if cantidad <= 0 and precio <= 0:
                continue
            actual[seccion].append({
                "codigo": "", "descripcion": b, "unidad": unidad,
                "cantidad": cantidad, "precio": precio})

    if actual and (actual["materiales"] or actual["mano_obra"]
                   or actual["equipo"]):
        apus.append(actual)
    return apus


import re  # noqa: E402

# Inicios de descripcion que son lineas de calculo/totales, no recursos.
# Se comparan con startswith (no subcadena) para no descartar materiales
# legitimos como 'SEÑAL PREVENTIVA' (que contiene 'iva').
_EXCLUIR_FLEX = (
    "total", "subtotal", "sub total", "cargas sociales", "beneficios sociales",
    "impuesto", "iva mano", "i.t", "herramientas =", "herramientas -",
    "herramientas menor", "mano de obra indir", "gastos generales", "utilidad",
    "descripcion", "insumo", "parametro", "precio unitario", "costo directo",
    "costos indirectos",
)
# Sinonimos de encabezado de columna.
_H_DESC = ("descripcion", "insumo", "detalle", "parametro")
_H_UND = ("unidad", "und", "unid")
_H_CANT = ("cantidad", "cant")
_H_PRECIO = ("precio prod", "precio unit", "precio", "unit.", "unit ", "p.unit",
             "unitario")


def _es_encabezado(norm: list[str]) -> bool:
    tiene_desc = any(any(n.startswith(h) for h in _H_DESC) for n in norm)
    tiene_und = any(any(n.startswith(h) for h in _H_UND) for n in norm)
    return tiene_desc and tiene_und


def _mapear_columnas(norm: list[str]):
    col_und = col_cant = col_precio = None
    for j, n in enumerate(norm):
        if col_und is None and any(n.startswith(h) for h in _H_UND):
            col_und = j
        elif col_cant is None and any(n.startswith(h) for h in _H_CANT):
            col_cant = j
        elif col_precio is None and any(h in n for h in _H_PRECIO):
            col_precio = j
    col_desc = (col_und - 1) if col_und else None
    return col_desc, col_und, col_cant, col_precio


def _sin_numeracion(texto: str) -> str:
    """Quita la numeracion/codigo inicial: '1.- MATERIALES' / 'A MATERIAL'."""
    return re.sub(r"^[\d\.\)\-\s>]+", "", texto).strip()


def _seccion_flex(texto: str) -> str | None:
    t = normalizar(_sin_numeracion(texto))
    if t.startswith("material"):
        return "materiales"
    if t.startswith("mano de obra") or t.startswith("obrero"):
        return "mano_obra"
    if t.startswith("equipo") or t.startswith("maquinaria"):
        return "equipo"
    return None


def _valor_etiqueta_inline(texto: str) -> str:
    """De 'Item: 1.1 Hormigon' o 'Actividad : X' devuelve lo que sigue a ':'."""
    val = texto.split(":", 1)[1] if ":" in texto else texto
    return " ".join(val.split()).strip()


def _valor_a_la_derecha(cols: list[str], idx: int) -> str:
    for j in range(idx + 1, len(cols)):
        if cols[j]:
            return cols[j]
    return ""


def _es_etiqueta(n: str, claves) -> bool:
    return any(re.match(rf"{c}\s*:?", n) for c in claves)


def _extraer_flexible(filas: list) -> list[dict]:
    """Reconoce Formularios B-2 de varios programas: detecta la actividad por
    'Item:'/'Actividad:', mapea las columnas desde la fila de encabezado
    (Descripcion/Insumo, Unidad/Und, Cantidad/Cant, Precio/Unit) y separa las
    secciones (1/A=Materiales, 2/B=Mano de obra/Obrero, 3/C=Equipo). Descarta
    totales, subtotales y lineas con '%'."""
    apus: list[dict] = []
    actual = None
    seccion = None
    col_desc = col_und = col_cant = col_precio = None

    def _nuevo(nombre):
        return {"actividad": nombre or "Actividad", "unidad": "",
                "cantidad": 1.0, "materiales": [], "mano_obra": [], "equipo": []}

    for row in filas:
        cols = [_txt(c) for c in row]
        norm = [normalizar(c) for c in cols]

        # 1) Fila de encabezado de columnas -> mapear posiciones.
        if _es_encabezado(norm):
            col_desc, col_und, col_cant, col_precio = _mapear_columnas(norm)
            continue

        # 2) Etiquetas de cabecera del APU (item/actividad, unidad, cantidad).
        etiqueta = False
        for j, n in enumerate(norm):
            if not n:
                continue
            if _es_etiqueta(n, ("item", "actividad")):
                if actual and (actual["materiales"] or actual["mano_obra"]
                               or actual["equipo"]):
                    apus.append(actual)
                nombre = (_valor_etiqueta_inline(cols[j])
                          or _valor_a_la_derecha(cols, j))
                actual = _nuevo(nombre)
                seccion = None
                col_desc = col_und = col_cant = col_precio = None
                # En la misma fila puede venir la unidad o 'cant unidad'.
                for k in range(j + 1, len(cols)):
                    cand = cols[k]
                    m = re.match(r"^\s*([\d.,]+)\s+([^\d\s].*)$", cand)
                    if m:
                        actual["cantidad"] = _num(m.group(1)) or 1.0
                        actual["unidad"] = m.group(2).strip()
                        break
                    if normalizar(cand).startswith(("unidad", "unid")):
                        actual["unidad"] = _valor_etiqueta_inline(cand)
                        break
                etiqueta = True
                break
            if _es_etiqueta(n, ("unitario", "unidades", "unidad")):
                if actual:
                    actual["unidad"] = (_valor_etiqueta_inline(cols[j])
                                        or _valor_a_la_derecha(cols, j))
                etiqueta = True
                break
            if _es_etiqueta(n, ("cantidad",)):
                if actual:
                    val = (_valor_etiqueta_inline(cols[j])
                           or _valor_a_la_derecha(cols, j))
                    actual["cantidad"] = _num(val) or 1.0
                etiqueta = True
                break
        if etiqueta:
            continue
        if actual is None:
            continue

        tiene_pct = any("%" in c for c in cols)

        def _num_col(idx):
            return (_num(cols[idx]) if idx is not None and idx < len(cols)
                    else 0.0)

        # 3) Cambio de seccion: fila con palabra clave y SIN cantidad/precio.
        sec_detectada = None
        for celda in cols:
            sec_detectada = _seccion_flex(celda)
            if sec_detectada:
                break
        if (sec_detectada and _num_col(col_cant) <= 0
                and _num_col(col_precio) <= 0):
            seccion = sec_detectada
            continue

        # 4) Fila de recurso segun las columnas detectadas.
        if seccion and col_desc is not None and col_desc < len(cols) \
                and not tiene_pct:
            desc = cols[col_desc]
            dn = normalizar(desc)
            if desc and not any(dn.startswith(k) for k in _EXCLUIR_FLEX):
                und = cols[col_und] if col_und is not None and \
                    col_und < len(cols) else ""
                cant = _num_col(col_cant)
                precio = _num_col(col_precio)
                if cant > 0 or precio > 0:
                    actual[seccion].append({
                        "codigo": "", "descripcion": desc, "unidad": und,
                        "cantidad": cant, "precio": precio})

    if actual and (actual["materiales"] or actual["mano_obra"]
                   or actual["equipo"]):
        apus.append(actual)
    return apus


def guardar_banco(apus: list[dict], proyecto: str = "",
                  reemplazar: bool = True) -> Path:
    """Guarda los APUs en el banco. Si reemplazar=False, los AGREGA a lo existente
    (evitando duplicados por nombre de actividad)."""
    ruta = settings.DATA_DIR / "banco_apu.json"
    existentes: list[dict] = []
    if not reemplazar and ruta.exists():
        try:
            existentes = json.loads(ruta.read_text(encoding="utf-8")).get(
                "apus", [])
        except Exception:
            existentes = []
    vistos = {normalizar(a.get("actividad", "")) for a in existentes}
    for a in apus:
        clave = normalizar(a.get("actividad", ""))
        if clave and clave not in vistos:
            existentes.append(a)
            vistos.add(clave)
    banco = {"_descripcion": "Banco de APU de referencia (rendimientos y precios "
             "reales). Usado por el motor y la IA.", "proyecto": proyecto,
             "apus": existentes}
    ruta.write_text(json.dumps(banco, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    return ruta


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python -m scripts.importar_apu_banco <ruta_excel>")
        return
    ruta = sys.argv[1]
    proyecto = ""
    if "--proyecto" in sys.argv:
        proyecto = sys.argv[sys.argv.index("--proyecto") + 1]
    apus = importar(ruta)
    salida = guardar_banco(apus, proyecto)
    print(f"Importados {len(apus)} APUs -> {salida}")
    # resumen
    for a in apus[:5]:
        print(f"  - {a['actividad'][:50]} [{a['unidad']}] "
              f"M:{len(a['materiales'])} MO:{len(a['mano_obra'])} "
              f"EQ:{len(a['equipo'])}")
    if len(apus) > 5:
        print(f"  ... y {len(apus) - 5} mas")


if __name__ == "__main__":
    main()
